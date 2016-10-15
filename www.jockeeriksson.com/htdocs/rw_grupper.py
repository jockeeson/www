#!/usr/bin/env python

from openpyxl import load_workbook
import itertools
import warnings

warnings.simplefilter("ignore")

wb = load_workbook(filename='./RTCC_EM-tips_2016.xlsm', read_only=True, data_only=True)
ws = wb.get_sheet_by_name('Grupper')

target = open('./grupper/grupper.html','w')
target.truncate()

s = ""
for j in range(0,6):
   s += '<h2>' + ws.cell(row = 3+j*9, column = 2).value + '</h2> \n'
   s += '<div>\n   <table>\n'
   s += '      <tr>\n'
   for i in [2,3,8,9,10,11,20]:
      if i == 3:
         s += '         <th class=team>' + ws.cell(row = 6+j*9, column = i).value + '</th> \n'
      elif i == 2:
         s += '         <th class=pos>' + str(ws.cell(row = 6+j*9, column = i).value) + '</th> \n'
      else:
         s += '         <th>' + str(ws.cell(row = 6+j*9, column = i).value) + '</th> \n'
   s += '      </tr>\n'
   
   #print type(ws.iter_rows('B15:D19'))
   for i in range(7+j*9,11+j*9):
      s += '      <tr>\n'
      s += '         <td class=pos>' + str(ws.cell(row = i, column = 2).value) + '</td> \n'
      s += '         <td class=team>' + ws.cell(row = i, column = 3).value + '</td> \n'
      s += '         <td>' + str(ws.cell(row = i, column = 8).value) + '</td> \n'
      s += '         <td>' + str(ws.cell(row = i, column = 9).value) + '</td> \n'
      s += '         <td>' + str(ws.cell(row = i, column = 10).value) + '</td> \n'
      s += '         <td>' + str(ws.cell(row = i, column = 11).value) + '</td> \n'
      s += '         <td>' + str(ws.cell(row = i, column = 20).value) + '</td> \n'
      s += '      </tr>\n'
   s += '   </table>\n</div>\n\n\n'


target.write(s.encode("utf-8"))

target.close()

