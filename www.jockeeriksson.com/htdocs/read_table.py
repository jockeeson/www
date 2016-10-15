#!/usr/bin/env python

from openpyxl import load_workbook
import warnings
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt


warnings.simplefilter("ignore")

wb = load_workbook(filename='./RTCC_EM-tips_2016.xlsm', read_only=True, data_only=True)
ws = wb.get_sheet_by_name('Sortering')

colors = ['blue', 'green', 'red', 'cyan', 'magenta', 'black']

legend = []
score_max = []
pos_max = []

score_fig = plt.figure()
score_ax = score_fig.add_subplot(111)
pos_fig = plt.figure()
pos_ax = pos_fig.add_subplot(111)

for j in range(1,wb.get_sheet_by_name('Resultat').cell(row=1,column=12).value+1):
   score = []
   pos = []
   games = []
   legend.append(ws.cell(row=j+6, column=2).value)
   for i in range(3,45):
      if ws.cell(row=j+6, column=i).value != "#N/A":
         score.append(ws.cell(row=j+6, column=i).value)
         pos.append(ws.cell(row=j+22, column=i).value)
         games.append(i-3)
      else:
         score_max.append(score[-1])
         pos_max.append(pos[-1])
         break
   
   score_ax.plot(games,score, label=legend[j-1], color=colors[j-1])
   score_ax.scatter(games,score, 50, color =colors[j-1])

   pos_ax.scatter(games,pos, 50, color =colors[j-1])
   pos_ax.plot(games,pos, label=legend[j-1], color=colors[j-1])

results = []
for i in range(3,45):
   if ws.cell(row=j+6, column=i).value != "#N/A":
      results.append(ws.cell(row=5, column=i).value)
   else:
      results.append('None')
      break

results[0] = ''

xmax = max(games)+1
ymax = max(score_max)+1

score_ax.set_xlim(0, xmax)
#score_ax.set_xlim(0.05, xmax+1)
score_ax.set_xticks(np.arange(1, xmax+1, 1.0))
score_ax.set_xticklabels(results[1:-1], rotation=-90)

score_ax.set_ylim(0, ymax)

score_ax.legend(loc='upper left', frameon=True)
score_ax.legend(bbox_to_anchor=(1.3, 0.8))

score_fig.subplots_adjust(top=0.95)
score_fig.subplots_adjust(left=0.05)
score_fig.subplots_adjust(bottom=0.4)
score_fig.subplots_adjust(right=0.75)
score_fig.savefig('./stats/points.pdf')


xmax = max(games)
ymax = max(pos_max)

pos_ax.set_xlim(0.05, xmax+1)
pos_ax.set_xticks(np.arange(1, xmax+1, 1.0))
pos_ax.set_xticklabels(results[1:-1], rotation=-90)

pos_ax.set_ylim(1-0.05, ymax+0.05)
pos_ax.invert_yaxis()
pos_ax.set_yticks(np.arange(1, ymax+1, 1.0))

pos_ax.legend(loc='upper left', frameon=True)
pos_ax.legend(bbox_to_anchor=(1.3, 0.8))

pos_fig.subplots_adjust(top=0.95)
pos_fig.subplots_adjust(left=0.05)
pos_fig.subplots_adjust(bottom=0.4)
pos_fig.subplots_adjust(right=0.75)
pos_fig.savefig('./stats/position.pdf')


