#!/usr/bin/env python

from openpyxl import load_workbook
import itertools
import warnings
import time

warnings.simplefilter("ignore")

wb = load_workbook(filename='./RTCC_EM-tips_2016.xlsm', read_only=True, data_only=True)
ws = wb.get_sheet_by_name('Resultat')

target = open('./matcher/matcher.html','w')
target.truncate()

s = ""
s += '<h2>Matchresultat</h2> \n'
s += '<div>\n   <table>\n'
s += '      <tr>\n'
s += '         <th class=match>Match</th> \n'
s += '         <th class=group>Grupp</th> \n'
s += '         <th class=results colspan="3">Resultat</th> \n'
s += '         <th class=date colspan="3">Datum</th> \n'
s += '      </tr>\n'

for j in range(13,64):
   s += '      <tr>\n'
   for i in [1,4,5,8,10,7,2]:
      if i is 1:
         s += '         <td class=match>' + str(ws.cell(row = j, column = i).value) + '</td> \n'
      elif i is 4:
         if type(ws.cell(row = j, column = i).value) is float:
            if ws.cell(row = j, column = i).value == 0.125:
               s += '         <td class=group>1/8</td> \n'
            elif ws.cell(row = j, column = i).value == 0.25:
               s += '         <td class=group>1/4</td> \n'
            #s += '         <td class=group>' + str(ws.cell(row = j, column = i).value) + '</td> \n'
         else:
            s += '         <td class=group>' + ws.cell(row = j, column = i).value + '</td> \n'
      elif i is 5:
         if ws.cell(row = j, column = i).value == None:
            s += '         <td class=results_h>'
         else:
            s += '         <td class=results_h>' + ws.cell(row = j, column = i).value + '  '
      elif i is 8:
         if ws.cell(row = j, column = i).value == None:
            s += '</td>\n         <td> - </td>\n'
         else:
            s += str(ws.cell(row = j, column = i).value) + '</td>\n         <td> - </td>'
      elif i is 10:
         if ws.cell(row = j, column = i).value == None:
            s += '         <td class=results_a>  '
         else:
            s +='         <td class=results_a>' +  str(ws.cell(row = j, column = i).value) + '  '
      elif i is 7:
         if ws.cell(row = j, column = i).value == None:
            s += '</td> \n'
         else:
            s += ws.cell(row = j, column = i).value + '</td> \n'
      elif i is 2:
         s += '         <td class=date>' + str(ws.cell(row = j, column = i).value.date()) + '</td>'
   s += '      </tr>\n'
s += '   </table>\n</div>\n\n\n'

target.write(s.encode("utf-8"))

target.close()

